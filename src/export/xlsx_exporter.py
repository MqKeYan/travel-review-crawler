"""
模块名称：XLSX 格式导出器

功能说明：
    - 将评论数据导出为 Excel XLSX 格式
    - 支持自定义字段选择
    - 自动调整列宽
    - 表头使用中文显示

依赖模块：
    - openpyxl (第三方)
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XlImage

from src.export.base import BaseExporter
from src.models.review import STANDARD_FIELDS
from src.utils.exceptions import ExportError


class XlsxExporter(BaseExporter):
    """Excel XLSX 格式导出器"""

    format_name: str = "xlsx"
    file_extension: str = "xlsx"

    def export(
        self,
        reviews: list[dict],
        filepath: str,
        fields: list[str] | None = None,
    ) -> str:
        """
        导出为 XLSX 格式。

        包含功能：
        - 自动列宽（根据内容长度）
        - 表头加粗、深色背景、白色文字
        - 内容区域自动换行

        Args:
            reviews: 评论数据列表
            filepath: 输出路径（不含扩展名）
            fields: 需要导出的字段列，None 则导出所有标准字段

        Returns:
            写入的文件完整路径

        Raises:
            ExportError: 写入文件失败
        """
        filepath = self._ensure_extension(filepath)

        # 确定字段和表头
        if fields:
            field_names = fields
            headers = self._get_headers(fields)
        else:
            field_names = [k for k, _ in STANDARD_FIELDS]
            headers = [v for _, v in STANDARD_FIELDS]

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "评论数据"

            # 定义样式
            header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2a2f2a", end_color="2a2f2a", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_alignment = Alignment(vertical="top", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

            # 写入表头
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            image_fields = {"image_urls", "avatar_url"}

            for row_idx, review in enumerate(reviews, 2):
                for col_idx, field in enumerate(field_names, 1):
                    value = review.get(field, "")

                    if field in image_fields and value:
                        # 图片/头像列：嵌入图片到单元格
                        paths = value if isinstance(value, list) else [value]
                        self._write_image_cell(ws, row_idx, col_idx, paths)
                    elif isinstance(value, list):
                        value = "; ".join(str(v) for v in value)
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                    else:
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.alignment = cell_alignment
                        cell.border = thin_border

            # 自动调整列宽（取表头和内容最大长度，限制最大宽度 50）
            for col_idx in range(1, len(headers) + 1):
                max_length = len(str(headers[col_idx - 1]))
                for row_idx in range(2, min(len(reviews) + 2, 200)):  # 样本前 200 行
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                    max_length = max(max_length, len(cell_value))
                # 限制列宽在 10~50 之间
                ws.column_dimensions[chr(64 + col_idx)].width = min(max(max_length * 1.2, 10), 50)

            # 冻结首行
            ws.freeze_panes = "A2"

            wb.save(filepath)

        except OSError as e:
            raise ExportError(f"XLSX 文件写入失败: {e}") from e
        except Exception as e:
            raise ExportError(f"XLSX 导出异常: {e}") from e

        return filepath

    def _get_headers(self, fields: list[str]) -> list[str]:
        """
        根据字段名获取对应的中文表头。

        Args:
            fields: 字段名列表

        Returns:
            中文表头列表
        """
        field_to_header = {k: v for k, v in STANDARD_FIELDS}
        return [field_to_header.get(f, f) for f in fields]

    def _write_image_cell(self, ws, row: int, col: int, image_paths: list[str]) -> None:
        """
        在 Excel 单元格中嵌入图片。

        对于本地存在的图片文件，嵌入并缩放到合适尺寸；
        对于 URL 链接或缺失文件，写入提示文本。

        Args:
            ws: openpyxl 工作表对象
            row: 行号（从 1 开始）
            col: 列号（从 1 开始）
            image_paths: 图片路径列表
        """
        valid_paths = [p for p in image_paths if isinstance(p, str) and os.path.isfile(p)]
        texts = []

        for p in image_paths:
            if not isinstance(p, str) or not p:
                continue
            if not os.path.isfile(p):
                texts.append(os.path.basename(p) if not p.startswith("http") else "[链接]")

        # 写入文本说明（在嵌入图片之前留底）
        if texts:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col)
            ws.cell(row=row, column=col, value="; ".join(texts))

        # 嵌入本地图片
        for i, path in enumerate(valid_paths):
            try:
                img = XlImage(path)
                # 缩放到合适尺寸（最大 120x100 像素）
                img.width = min(img.width, 120)
                img.height = min(img.height, 100)
                # 锚定到单元格
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col)
                anchor = f"{col_letter}{row}"
                ws.add_image(img, anchor)
            except Exception:
                pass
